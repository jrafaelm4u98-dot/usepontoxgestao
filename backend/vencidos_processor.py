"""
Processador de Boletos Vencidos Estrutural
==========================================
Cruza:
  - ESTRUTURAL LISTAGEM  (NOSSO_NUMERO, NOME_VENDEDOR, VALOR, COD_PDV, NOME_PDV)
  - BASE PDV M4U         (COD_PDV, NOME_PDV)  ← opcional, enriquece descrição
  - Paranaue DDD 42/47/61/63 (DATA_CADASTRO, VENCIMENTO, NOME_VENDEDOR, NOSSO_NUMERO)

Gera: Excel com abas 42 / 47 / 61 / 63
"""

import pandas as pd
import numpy as np
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Mapeamento flexível de colunas ─────────────────────────────────────────
# Normaliza nomes de coluna para facilitar o match, independente do cabeçalho exato da planilha.

def _norm(s: str) -> str:
    """Remove acentos, espaços extras e caracteres estranhos."""
    import unicodedata
    import re
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^a-z0-9]', '', s) # Remove tudo que não for letra ou número
    return s

def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Retorna o nome real da coluna que casa com um dos candidatos normalizados."""
    norms = {_norm(c): c for c in df.columns}
    for cand in candidates:
        if _norm(cand) in norms:
            return norms[_norm(cand)]
    return None

def _require_col(df: pd.DataFrame, candidates: list[str], label: str) -> str:
    col = _find_col(df, candidates)
    if col is None:
        raise ValueError(
            f"Coluna '{label}' não encontrada.\n"
            f"Esperava uma de: {candidates}\n"
            f"Colunas disponíveis: {list(df.columns)}"
        )
    return col


# ── Leitura genérica de planilhas ──────────────────────────────────────────

def _read_xlsx(file_like) -> pd.DataFrame:
    """Lê arquivo Excel ou XML Spreadsheet exportado por bancos disfarçado de .xls."""
    data = file_like.read() if hasattr(file_like, 'read') else file_like
    buf = BytesIO(data)
    
    # ── Tenta ler como XML Spreadsheet 2003 (Muitos sistemas bancários exportam assim) ──
    buf.seek(0)
    head_bytes = buf.read(500)
    if b'<?xml' in head_bytes and b'urn:schemas-microsoft-com:office:spreadsheet' in head_bytes:
        import html, re
        buf.seek(0)
        raw_bytes = buf.read()
        try:
            content = raw_bytes.decode('utf-8')
        except UnicodeDecodeError:
            content = raw_bytes.decode('iso-8859-1', errors='ignore')
            
        rows = re.findall(r'<Row[^>]*>(.*?)</Row>', content, re.DOTALL | re.IGNORECASE)
        if not rows:
            rows = re.findall(r'<ss:Row[^>]*>(.*?)</ss:Row>', content, re.DOTALL | re.IGNORECASE)
        
        data_matrix = []
        for row in rows:
            cells = []
            c_matches = re.findall(r'<Data[^>]*>(.*?)</Data>', row, re.IGNORECASE | re.DOTALL)
            if not c_matches:
                c_matches = re.findall(r'<ss:Data[^>]*>(.*?)</ss:Data>', row, re.IGNORECASE | re.DOTALL)
            for c in c_matches:
                clean_text = html.unescape(re.sub(r'<[^>]*>', '', c)).strip()
                cells.append(clean_text)
            data_matrix.append(cells)
        
        df = pd.DataFrame(data_matrix)
        df = df.dropna(how='all').reset_index(drop=True)
        if not df.empty:
            return _find_header(df)

    # ── Se não for XML, tenta pelas vias originais do Pandas ──
    for engine in ('openpyxl', 'xlrd', 'pyxlsb'):
        try:
            buf.seek(0)
            df = pd.read_excel(buf, engine=engine, header=None)
            df = df.dropna(how='all').reset_index(drop=True)
            if df.empty:
                return df
            return _find_header(df)
        except Exception:
            continue
    raise ValueError("Não foi possível ler o arquivo Excel.")

def _find_header(df: pd.DataFrame) -> pd.DataFrame:
    """Encontra a linha de cabeçalho verdadeira nas 15 primeiras linhas."""
    head_idx = 0
    max_non_null = 0
    for i, row in df.head(15).iterrows():
        valid_cells = sum(1 for val in row if pd.notna(val) and str(val).strip() != '')
        if valid_cells > max_non_null:
            max_non_null = valid_cells
            head_idx = i
            
    df.columns = [str(c).strip() for c in df.iloc[head_idx]]
    df = df.iloc[head_idx + 1:].reset_index(drop=True)
    return df


# ── Processamento principal ────────────────────────────────────────────────

def process_boletos_vencidos(
    estrutural_bytes: bytes,
    base_pdv_bytes: bytes | None,
    paranaue_map: dict[str, bytes],   # {'42': bytes, '47': bytes, ...}
) -> bytes:
    """
    Retorna bytes de um Excel com abas por DDD.
    paranaue_map: chaves são os DDDs como string ('42', '47', '61', '63').
    """

    # 1. Carrega ESTRUTURAL LISTAGEM
    df_est = _read_xlsx(BytesIO(estrutural_bytes))

    col_nn_est   = _require_col(df_est, ['nosso numero', 'nosso_numero', 'nossonumero', 'num', 'numero'], 'Nosso Número (Estrutural)')
    col_vend_est = _find_col(df_est,    ['nome vendedor', 'vendedor', 'colaborador', 'nome do vendedor'])
    col_venc_est = _find_col(df_est,    ['vencimento', 'data vencimento', 'dt vencimento', 'data venc', 'venc'])
    col_valor    = _require_col(df_est, ['valor', 'valor r$', 'valor rs', 'vlr', 'vl', 'valor deposito'], 'Valor (Estrutural)')
    col_cod_pdv  = _require_col(df_est, ['codigo pdv', 'cod pdv', 'codigo', 'cod', 'cod. pdv', 'codigo do pdv'], 'Código PDV')
    col_nome_pdv = _require_col(df_est, ['nome do pdv', 'nome pdv', 'descricao', 'descr', 'pdv', 'estabelecimento'], 'Nome PDV')

    # Normaliza NOSSO NUMERO removendo .0 e arrancando ZEROS à esquerda para garantir MATCH
    df_est[col_nn_est] = df_est[col_nn_est].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')

    # 2. Carrega BASE PDV (opcional) para enriquecer NOME DO PDV
    df_pdv = None
    if base_pdv_bytes:
        try:
            df_pdv = _read_xlsx(BytesIO(base_pdv_bytes))
            _bc = _find_col(df_pdv, ['codigo pdv', 'cod pdv', 'codigo', 'cod', 'cod pdv.'])
            _bn = _find_col(df_pdv, ['nome do pdv', 'nome pdv', 'descricao', 'pdv', 'nome', 'razao'])
            if _bc and _bn:
                df_pdv = df_pdv[[_bc, _bn]].copy()
                df_pdv.columns = ['__cod_pdv', '__nome_pdv']
                # Retira nulos, acerta formato e arranca zeros à esquerda
                df_pdv['__cod_pdv'] = df_pdv['__cod_pdv'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')
                df_pdv = df_pdv.dropna(subset=['__cod_pdv'])
                df_pdv = df_pdv.drop_duplicates(subset=['__cod_pdv'], keep='first')
            else:
                df_pdv = None
        except Exception:
            df_pdv = None

    # 3. Processa cada DDD
    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    ddds_processados = []

    for ddd, par_bytes in paranaue_map.items():
        try:
            df_par = _read_xlsx(BytesIO(par_bytes))
        except Exception as e:
            raise ValueError(f"Erro ao ler planilha DDD {ddd}: {e}")

        # Colunas do Paranaue (DATA DEPÓSITO também serve como DATA CADASTRO)
        col_nn_par   = _require_col(df_par, ['nosso numero', 'nosso_numero', 'nossonumero', 'num', 'numero'], f'Nosso Número (DDD {ddd})')
        col_cad      = _find_col(df_par,    ['data cadastro', 'data do cadastro', 'dt cadastro', 'cadastro', 'data cad', 'data deposito', 'deposito'])
        col_vend_par = _find_col(df_par,    ['nome vendedor', 'vendedor', 'colaborador', 'nome do vendedor', 'nome colaborador'])

        # Normaliza NN arrancando ZEROS à esquerda
        df_par[col_nn_par] = df_par[col_nn_par].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')

        # Merge: paranaue LEFT JOIN estrutural para trazer VALOR + PDV + VENCIMENTO
        lookup_cols = [col_nn_est, col_valor, col_cod_pdv, col_nome_pdv]
        if col_vend_est: lookup_cols.append(col_vend_est)
        if col_venc_est: lookup_cols.append(col_venc_est)

        lookup = df_est[lookup_cols].copy()
        lookup = lookup.rename(columns={col_nn_est: '__nn'})
        
        # Remove eventuais duplicações no NOSSO NÚMERO da Estrutural para evitar multiplicar linhas do Paranaue
        lookup = lookup.drop_duplicates(subset=['__nn'], keep='first')

        df_par_m = df_par.copy()
        # '__nn' em lookup também já teve os zeros arrancados na fase 1!
        df_par_m = df_par_m.merge(lookup, left_on=col_nn_par, right_on='__nn', how='left')
        df_par_m.drop(columns=['__nn'], errors='ignore', inplace=True)

        # Enriquece NOME PDV com BASE PDV se disponível
        if df_pdv is not None:
            # Também arranca zeros à esquerda no Código do Boleto para garantir o match exato com a Base
            df_par_m[col_cod_pdv] = df_par_m[col_cod_pdv].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).str.lstrip('0')
            df_par_m = df_par_m.merge(df_pdv, left_on=col_cod_pdv, right_on='__cod_pdv', how='left')
            # usa nome do PDV da base quando disponível e recusa se a base devolver 'nan' textual
            df_par_m['__nome_pdv'] = df_par_m['__nome_pdv'].replace('nan', np.nan)
            df_par_m[col_nome_pdv] = df_par_m['__nome_pdv'].fillna(df_par_m[col_nome_pdv])
            df_par_m.drop(columns=['__cod_pdv', '__nome_pdv'], errors='ignore', inplace=True)

        # Decide NOME COLABORADOR
        if col_vend_par and col_vend_par in df_par_m.columns:
            nome_col_src = col_vend_par
        elif col_vend_est and col_vend_est in df_par_m.columns:
            nome_col_src = col_vend_est
        else:
            nome_col_src = None

        # Monta linha de saída
        def _get(col, fallback=''):
            return df_par_m[col].fillna(fallback).replace('nan', fallback) if col and col in df_par_m.columns else fallback

        df_out = pd.DataFrame()
        df_out['DATA CADASTRO']    = _get(col_cad)
        df_out['VENCIMENTO']       = _get(col_venc_est)    # Agora o vencimento vem da estrutural!
        df_out['NOME COLABORADOR'] = _get(nome_col_src)
        
        # Reconstrói os zeros do NOSSO NÚMERO (mantém até 10 dígitos)
        df_out['NOSSO NUMERO']     = df_par_m[col_nn_par].apply(lambda x: str(x).zfill(10) if x and x != 'nan' else '')
        
        df_out['VALOR R$']         = _get(col_valor)
        
        codigo_series = _get(col_cod_pdv)
        desc_series = _get(col_nome_pdv)
        
        # Se 'codigo' for '-', nulo ou vazio, define a descrição como 'BOLETO VENDEDOR'
        def calc_desc(row):
            cod, desc = str(row['__cod']).strip(), str(row['__desc']).strip()
            if not cod or cod == '-' or cod == 'nan':
                return 'BOLETO VENDEDOR'
            return desc
            
        temp_df = pd.DataFrame({'__cod': codigo_series, '__desc': desc_series})
        df_out['CÓDIGO']           = temp_df['__cod']
        df_out['DESCRIÇÃO']        = temp_df.apply(calc_desc, axis=1)

        # Ordena por VENCIMENTO se disponível
        if col_venc_est and col_venc_est in df_par_m.columns:
            try:
                # Criar coluna temporária segura para ordenar sem destruir o texto original caso a conversão falhe
                temp_date = pd.to_datetime(df_out['VENCIMENTO'], dayfirst=True, errors='coerce')
                # Define a ordem baseada na data real
                df_out = df_out.iloc[temp_date.argsort()]
                
                # E também formata o que for data válida, preservando o que não for
                def safe_parse(x):
                    if pd.isna(x) or str(x).strip() in ('', 'nan'): return ''
                    if isinstance(x, pd.Timestamp): return x
                    try: return pd.to_datetime(x, dayfirst=True)
                    except: return x
                
                df_out['VENCIMENTO'] = df_out['VENCIMENTO'].apply(safe_parse)
                df_out['DATA CADASTRO'] = df_out['DATA CADASTRO'].apply(safe_parse)
            except Exception:
                pass

        # Gera aba no Excel
        _write_sheet(wb, f'DDD {ddd}', df_out)
        ddds_processados.append(ddd)

    if not ddds_processados:
        raise ValueError("Nenhum arquivo do Paranaue foi enviado.")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Geração de aba formatada ───────────────────────────────────────────────

_HEADER_FILL  = PatternFill('solid', fgColor='1E293B')
_HEADER_FONT  = Font(bold=True, color='F1F5F9', size=10)
_ALT_FILL     = PatternFill('solid', fgColor='F8FAFC')
_BORDER_SIDE  = Side(style='thin', color='E2E8F0')
_CELL_BORDER  = Border(left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE)
_DATE_FMT     = 'DD/MM/YYYY'
_MONEY_FMT    = 'R$ #,##0.00'

_COL_WIDTHS = {
    'DATA CADASTRO':    14,
    'VENCIMENTO':       14,
    'NOME COLABORADOR': 28,
    'NOSSO NUMERO':     16,
    'VALOR R$':         14,
    'CÓDIGO':           14,
    'DESCRIÇÃO':        36,
}

def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(title=sheet_name)

    # Cabeçalho
    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill       = _HEADER_FILL
        cell.font       = _HEADER_FONT
        cell.alignment  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border     = _CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = _COL_WIDTHS.get(col_name, 18)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    # Linhas de dados
    for ri, row in enumerate(df.itertuples(index=False), 2):
        alt = (ri % 2 == 0)
        for ci, val in enumerate(row, 1):
            col_name = df.columns[ci - 1]
            cell = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else '')

            # Formatos especiais
            if col_name in ('DATA CADASTRO', 'VENCIMENTO') and isinstance(val, pd.Timestamp):
                cell.value  = val.to_pydatetime()
                cell.number_format = _DATE_FMT
            elif col_name == 'VALOR R$':
                if isinstance(val, (int, float)):
                    cell.value = float(val)
                else:
                    try:
                        sv = str(val).replace('R$', '').replace('\xa0', '').strip()
                        # Formato 1.500,50
                        if ',' in sv and '.' in sv:
                            sv = sv.replace('.', '').replace(',', '.')
                        # Formato 1500,50
                        elif ',' in sv:
                            sv = sv.replace(',', '.')
                        # Se já for 1500.50 não quebra nada
                        cell.value = float(sv)
                    except Exception:
                        pass
                cell.number_format = _MONEY_FMT

            cell.alignment = Alignment(vertical='center', horizontal='center' if col_name in ('DATA CADASTRO', 'VENCIMENTO', 'NOSSO NUMERO', 'CÓDIGO', 'VALOR R$') else 'left')
            cell.border = _CELL_BORDER
            if alt:
                cell.fill = _ALT_FILL

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
