import pandas as pd
import numpy as np
import io
import re
import html
from thefuzz import fuzz
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import unicodedata

def robust_read_excel(buffer, engine):
    buffer.seek(0)
    head_bytes = buffer.read(500)
    buffer.seek(0)
    
    if b'<?xml' in head_bytes and b'urn:schemas-microsoft-com:office:spreadsheet' in head_bytes:
        # Pega a string ignorando erros de encoding
        content = buffer.read().decode('iso-8859-1', errors='ignore')
        rows = re.findall(r'<Row[^>]*>(.*?)</Row>', content, re.DOTALL | re.IGNORECASE)
        # Se os namespaces forem diferentes, tenta buscar com ss:Row
        if not rows:
            rows = re.findall(r'<ss:Row[^>]*>(.*?)</ss:Row>', content, re.DOTALL | re.IGNORECASE)
            
        data = []
        for row in rows:
            cells = []
            c_matches = re.findall(r'<Data[^>]*>(.*?)</Data>', row, re.IGNORECASE | re.DOTALL)
            if not c_matches:
                c_matches = re.findall(r'<ss:Data[^>]*>(.*?)</ss:Data>', row, re.IGNORECASE | re.DOTALL)
                
            for c in c_matches:
                # Remove tags internas <B> etc e desescapa HTML entities
                clean_text = html.unescape(re.sub(r'<[^>]*>', '', c)).strip()
                cells.append(clean_text)
            data.append(cells)
        return pd.DataFrame(data)
        
    return pd.read_excel(buffer, engine=engine, header=None)


def normalize_name(name):
    if pd.isna(name) or not str(name).strip():
        return ''
    n = str(name).lower().strip()
    # Normalize unicode to remove accents
    n = ''.join(c for c in unicodedata.normalize('NFD', n) if unicodedata.category(c) != 'Mn')
    n = re.sub(r'\s+', ' ', n)
    return n

def vendors_match_fuzz(name1, name2):
    n1 = normalize_name(name1)
    n2 = normalize_name(name2)
    if not n1 or not n2: return False
    if n1 == n2: return True
    if n1 in n2 or n2 in n1: return True
    
    # First + Last name exact check
    w1 = n1.split()
    w2 = n2.split()
    if len(w1) >= 2 and len(w2) >= 2:
        if w1[0] == w2[0] and w1[-1] == w2[-1]:
            return True
            
    # Levenshtein fuzz distance (simulating allowedEdits 0, 1, or 2 from original logic)
    # thefuzz uses ratio 0-100. Let's use distance instead or stick to high ratio.
    # To mirror allowedEdits (distance <= 2 for length > 10, distance <= 1 for 5..10, exact < 5)
    from Levenshtein import distance
    dist = distance(n1, n2)
    max_len = max(len(n1), len(n2))
    allowed = 0
    if max_len > 10: allowed = 2
    elif max_len >= 5: allowed = 1
    
    return dist <= allowed

def parse_sales(file_obj, filename, ddd):
    file_obj.seek(0)
    buffer = io.BytesIO(file_obj.read())
    file_lower = str(filename).lower()
    engine = 'xlrd' if file_lower.endswith('.xls') else ('pyxlsb' if file_lower.endswith('.xlsb') else 'openpyxl')
    df_raw = robust_read_excel(buffer, engine)
    sales = []
    
    # Skips row 0 (header) and empty rows
    for i in range(1, len(df_raw)):
        row = df_raw.iloc[i]
        
        vendor = row[0] if len(row) > 0 else np.nan
        if pd.isna(vendor) or not str(vendor).strip():
            continue
            
        vendor_str = str(vendor).strip()
        if 'total' in vendor_str.lower():
            continue
            
        try:
            val_liq_vendas = float(row[2]) if not pd.isna(row[2]) else 0.0
        except (ValueError, TypeError): val_liq_vendas = 0.0
            
        try:
            val_liq_cob = float(row[5]) if len(row) > 5 and not pd.isna(row[5]) else 0.0
        except (ValueError, TypeError): val_liq_cob = 0.0
            
        sales.append({
            'vendor': vendor_str,
            'ddd': int(ddd),
            'valor_liquido_vendas': val_liq_vendas,
            'valor_liquido_cobrancas': val_liq_cob
        })
    return pd.DataFrame(sales)

def group_sales(df_sales):
    if df_sales.empty:
        return []
        
    grouped = {}
    for idx, row in df_sales.iterrows():
        norm = normalize_name(row['vendor'])
        if norm in grouped:
            grouped[norm]['valor_liquido_vendas'] += row['valor_liquido_vendas']
            grouped[norm]['valor_liquido_cobrancas'] += row['valor_liquido_cobrancas']
        else:
            grouped[norm] = row.to_dict()
    return list(grouped.values())

def parse_desmembramentos(file_obj, filename):
    file_obj.seek(0)
    buffer = io.BytesIO(file_obj.read())
    file_lower = str(filename).lower()
    engine = 'xlrd' if file_lower.endswith('.xls') else ('pyxlsb' if file_lower.endswith('.xlsb') else 'openpyxl')
    df_raw = robust_read_excel(buffer, engine)
    desm = []
    
    if len(df_raw) < 2: return desm
    
    header = df_raw.iloc[0].astype(str).str.lower().str.strip()
    col_map = {'filial': -1, 'vendor': -1, 'pdv': -1, 'value': -1, 'vencimento': -1}
    
    for i, col in enumerate(header):
        if 'filial' in col: col_map['filial'] = i
        elif 'vendedor' in col: col_map['vendor'] = i
        elif 'pdv' in col or 'código' in col: col_map['pdv'] = i
        elif 'valor' in col: col_map['value'] = i
        elif 'vencimento' in col: col_map['vencimento'] = i
        
    if col_map['filial'] == -1: col_map['filial'] = 2
    if col_map['vendor'] == -1: col_map['vendor'] = 3
    if col_map['pdv'] == -1: col_map['pdv'] = 4
    if col_map['value'] == -1: col_map['value'] = 5
    if col_map['vencimento'] == -1: col_map['vencimento'] = 7
    
    for i in range(1, len(df_raw)):
        row = df_raw.iloc[i]
        
        try: filial = str(row[col_map['filial']]) if len(row) > col_map['filial'] else ""
        except: filial = ""
        
        try: vendor = str(row[col_map['vendor']]) if len(row) > col_map['vendor'] else ""
        except: vendor = ""
        
        try: pdv = str(row[col_map['pdv']]) if len(row) > col_map['pdv'] else ""
        except: pdv = ""
        
        try: raw_val = row[col_map['value']]
        except: raw_val = 0
        
        # ddd logic
        m = re.search(r'\d{2}', filial)
        ddd = int(m.group(0)) if m else None
        
        # Parse value
        if pd.isna(raw_val) or raw_val == "":
            val = 0.0
        elif isinstance(raw_val, (int, float)):
            val = float(raw_val)
        else:
            val_str = str(raw_val).replace('.', '').replace(',', '.')
            try: val = float(val_str)
            except: val = 0.0
            
        if not vendor or not pdv or pd.isna(vendor) or val <= 0:
            continue
            
        # Vencimento date hack
        try: raw_venc = row[col_map['vencimento']]
        except: raw_venc = ""
        
        if pd.isna(raw_venc):
            venc_str = ""
        elif isinstance(raw_venc, datetime):
            venc_str = raw_venc.strftime("%d/%m/%Y")
        else:
            venc_str = str(raw_venc).strip()
            # If it looks like 2026-04-14 00:00:00
            if len(venc_str) > 10 and '-' in venc_str:
                try: venc_str = datetime.strptime(venc_str[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                except: pass
                
        desm.append({
            'ddd': ddd,
            'vendor': vendor.strip(),
            'pdvCode': pdv.strip(),
            'value': val,
            'vencimento': venc_str
        })
        
    return desm

def group_desmembramentos(desm_list):
    grouped = {}
    for d in desm_list:
        norm = normalize_name(d['vendor'])
        key = f"{norm}_{d['pdvCode']}_{d['vencimento']}"
        if key in grouped:
            grouped[key]['value'] += d['value']
        else:
            grouped[key] = dict(d) # Copy
    return list(grouped.values())

def split_boleto(value, ddd, period_str):
    limit = 1000 if ddd == 63 else (5000 if ddd == 61 else float('inf'))
    if value <= limit:
        return [value]
        
    m = re.search(r'^(\d{2})', period_str)
    day = int(m.group(1)) if m else 1
    
    splits = []
    rem = value
    idx = 0
    while rem > 0 and idx <= 100:
        if ddd == 63:
            s_val = min(rem, 900 + day + (idx * 0.10))
        elif ddd == 61:
            s_val = min(rem, 4900 + day + (idx * 0.10))
        else:
            s_val = rem
        splits.append(min(s_val, rem))
        rem -= s_val
        idx += 1
    return splits

def process_business_rules(all_sales, all_desm, period):
    grouped_sales = group_sales(pd.DataFrame(all_sales))
    grouped_desm = group_desmembramentos(all_desm)
    
    # Apply DDD Rules
    for s in grouped_sales:
        ddd = s['ddd']
        if ddd in (42, 47):
            s['finalValue'] = s['valor_liquido_vendas'] + s['valor_liquido_cobrancas']
        elif ddd in (61, 63):
            s['finalValue'] = s['valor_liquido_vendas']
        else:
            s['finalValue'] = s['valor_liquido_vendas']
            
    # Apply Desmembramentos
    results = []
    for s in grouped_sales:
        # Match desm
        vendor_desm = [d for d in grouped_desm if d['ddd'] == s['ddd'] and vendors_match_fuzz(d['vendor'], s['vendor'])]
        
        if len(vendor_desm) > 0:
            total_desm = sum(d['value'] for d in vendor_desm)
            rem = s['finalValue'] - total_desm
            
            if rem > 0:
                results.append({
                    'vendor': s['vendor'],
                    'ddd': s['ddd'],
                    'finalValue': rem,
                    'isDesmembramento': False,
                    'pdvCode': None,
                    'vencimento': None
                })
                
            for d in vendor_desm:
                results.append({
                    'vendor': s['vendor'],
                    'ddd': s['ddd'],
                    'finalValue': d['value'],
                    'isDesmembramento': True,
                    'pdvCode': d['pdvCode'],
                    'vencimento': d['vencimento']
                })
        else:
            results.append({
                'vendor': s['vendor'],
                'ddd': s['ddd'],
                'finalValue': s['finalValue'],
                'isDesmembramento': False,
                'pdvCode': None,
                'vencimento': None
            })
            
    # Apply splits (limits DDD 61/63)
    final_data = []
    for r in results:
        splits = split_boleto(r['finalValue'], r['ddd'], period)
        if len(splits) == 1:
            final_data.append(r)
        else:
            for idx, s_val in enumerate(splits):
                new_r = dict(r)
                new_r['finalValue'] = s_val
                new_r['isSplit'] = True
                new_r['splitIndex'] = idx + 1
                final_data.append(new_r)
                
    return final_data

def generate_excel_bytes(data, period, starting_number):
    import io
    # Data is sorted/grouped by DDD
    data_by_ddd = {}
    for i, d in enumerate(data):
        d['numero'] = starting_number + i
        ddd = str(d.get('ddd', 'Outros'))
        if ddd not in data_by_ddd: data_by_ddd[ddd] = []
        data_by_ddd[ddd].append(d)
        
    output = io.BytesIO()
    import pandas as pd
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for ddd in sorted(data_by_ddd.keys()):
            sheet_name = f"DDD {ddd}"
            # Emulate the headers correctly
            # We don't use DataFrame direct to exact styling, but we can do it via writer.book
            df = pd.DataFrame(data_by_ddd[ddd])
            rows = []
            for _, row in df.iterrows():
                rows.append({
                    'Nº Número': row.get('numero', ''),
                    'Vendedor': row.get('vendor', ''),
                    'Valor R$': row.get('finalValue', 0.0),
                    'Vencimento': row.get('vencimento', '') or '',
                    'Código PDV': row.get('pdvCode', '') or ''
                })
                
            out_df = pd.DataFrame(rows)
            out_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False, header=False)
            
            # Now style the sheet using openpyxl
            sheet = writer.sheets[sheet_name]
            
            # TITLE (Row 1)
            sheet.merge_cells('A1:E1')
            title_cell = sheet['A1']
            title_cell.value = f"BOLETO ESTRUTURAL DDD {ddd} - PERÍODO {period}"
            title_cell.font = Font(bold=True, size=14, color="000000")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            sheet.row_dimensions[1].height = 30
            
            # HEADER (Row 2)
            headers = ['Nº Número', 'Vendedor', 'Valor R$', 'Vencimento', 'Código PDV']
            for col_idx, h in enumerate(headers, 1):
                cell = sheet.cell(row=2, column=col_idx)
                cell.value = h
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Border
                side = Side(border_style='thin', color='000000')
                cell.border = Border(left=side, right=side, top=side, bottom=side)
                
            # Content styling & coloring
            for row_idx, r in enumerate(data_by_ddd[ddd], start=3):
                row_cells = [sheet.cell(row=row_idx, column=col) for col in range(1, 6)]
                
                # Currency format
                row_cells[2].number_format = '"R$" #,##0.00'
                
                # Desmembramento Highlight
                if r.get('isDesmembramento') and r.get('pdvCode'):
                    row_cells[4].font = Font(color="FF0000", bold=True)
                    row_cells[1].font = Font(color="666666", italic=True)
                    
                # Alternate Colors and Borders
                is_even = ((row_idx - 3) % 2 == 0)
                fill_color = "F8F9FA" if is_even else "FFFFFF"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                side = Side(border_style='thin', color="CCCCCC")
                border = Border(left=side, right=side, top=side, bottom=side)
                
                for cell in row_cells:
                    if not cell.fill or getattr(cell.fill, 'fgColor', None) == '00000000':
                        cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            # Column Widths
            sheet.column_dimensions['A'].width = 12
            sheet.column_dimensions['B'].width = 35
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15

    output.seek(0)
    return output.read()
